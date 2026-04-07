"""
Validador Pp - Partida Específica
Aplicación para validar combinaciones de Programa presupuestario y Partida
según el catálogo oficial de SADER.
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN DE PÁGINA
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Validador Pp-Partida | SADER",
    page_icon="✓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ══════════════════════════════════════════════════════════════════════════════
# ESTILOS CSS PERSONALIZADOS
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
    /* Fuentes */
    @import url('https://fonts.googleapis.com/css2?family=Source+Sans+Pro:wght@400;600;700&display=swap');
    
    /* Colores institucionales SADER */
    :root {
        --guinda: #6B1D3D;
        --guinda-claro: #8B2D4D;
        --crema: #F5F0E6;
        --verde-ok: #2E7D32;
        --rojo-error: #C62828;
    }
    
    /* Header principal */
    .main-header {
        background: linear-gradient(135deg, var(--guinda) 0%, var(--guinda-claro) 100%);
        color: white;
        padding: 1.5rem 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 15px rgba(107, 29, 61, 0.3);
    }
    
    .main-header h1 {
        margin: 0;
        font-size: 1.8rem;
        font-weight: 700;
    }
    
    .main-header p {
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
        font-size: 1rem;
    }
    
    /* Tarjetas de estadísticas */
    .stat-card {
        background: white;
        border-radius: 10px;
        padding: 1.2rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        border-left: 4px solid var(--guinda);
        margin-bottom: 1rem;
    }
    
    .stat-card.success {
        border-left-color: var(--verde-ok);
    }
    
    .stat-card.error {
        border-left-color: var(--rojo-error);
    }
    
    .stat-number {
        font-size: 2rem;
        font-weight: 700;
        color: var(--guinda);
        line-height: 1;
    }
    
    .stat-label {
        color: #666;
        font-size: 0.9rem;
        margin-top: 0.3rem;
    }
    
    /* Resultado de validación individual */
    .result-valid {
        background: #E8F5E9;
        border: 1px solid #A5D6A7;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    
    .result-invalid {
        background: #FFEBEE;
        border: 1px solid #EF9A9A;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    
    /* Lista de partidas */
    .partida-list {
        background: var(--crema);
        border-radius: 8px;
        padding: 1rem;
        margin-top: 1rem;
    }
    
    .partida-chip {
        display: inline-block;
        background: white;
        border: 1px solid #ddd;
        border-radius: 4px;
        padding: 0.2rem 0.5rem;
        margin: 0.2rem;
        font-family: monospace;
        font-size: 0.85rem;
    }
    
    /* Sidebar */
    .sidebar .sidebar-content {
        background: var(--crema);
    }
    
    /* Botones */
    .stButton > button {
        background: var(--guinda);
        color: white;
        border: none;
        border-radius: 6px;
        padding: 0.5rem 1.5rem;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stButton > button:hover {
        background: var(--guinda-claro);
        box-shadow: 0 4px 12px rgba(107, 29, 61, 0.3);
    }
    
    /* Ocultar elementos de Streamlit */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* Ajustes responsivos */
    @media (max-width: 768px) {
        .main-header h1 {
            font-size: 1.4rem;
        }
    }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE PROCESAMIENTO
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data
def cargar_catalogo(archivo):
    """Carga y procesa el catálogo de Pp-Partidas."""
    df = pd.read_excel(archivo, header=None, dtype=str)
    df = df.iloc[1:].reset_index(drop=True)
    
    partidas_por_pp = {}
    
    for _, row in df.iterrows():
        mod = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
        prog = str(row.iloc[4]).strip().zfill(3) if pd.notna(row.iloc[4]) else ''
        partida = str(row.iloc[6]).strip().zfill(5) if pd.notna(row.iloc[6]) else ''
        
        if mod and prog and partida and partida != 'nan' and partida != '00nan':
            pp = f"{mod}{prog}"
            if pp not in partidas_por_pp:
                partidas_por_pp[pp] = set()
            partidas_por_pp[pp].add(partida)
    
    return partidas_por_pp


def procesar_archivo_validacion(archivo, partidas_por_pp):
    """Procesa archivo de claves a validar."""
    df_raw = pd.read_excel(archivo, header=None, dtype=str)
    datos = []
    
    # Buscar fila donde empiezan los datos (formato PIPP)
    fila_datos = None
    for i in range(min(15, len(df_raw))):
        val0 = str(df_raw.iloc[i, 0]).strip() if pd.notna(df_raw.iloc[i, 0]) else ''
        val1 = str(df_raw.iloc[i, 1]).strip() if df_raw.shape[1] > 1 and pd.notna(df_raw.iloc[i, 1]) else ''
        
        if val0.isdigit() and len(val0) <= 2 and int(val0) > 0:
            fila_datos = i
            break
        if val1.isdigit() and len(val1) <= 2 and int(val1) > 0:
            fila_datos = i
            break
    
    if fila_datos is not None:
        # Formato PIPP oficial
        df_datos = df_raw.iloc[fila_datos:].reset_index(drop=True)
        
        for _, row in df_datos.iterrows():
            pp_val = str(row.iloc[9]).strip().upper() if len(row) > 9 and pd.notna(row.iloc[9]) else ''
            partida_val = str(row.iloc[10]).strip().zfill(5) if len(row) > 10 and pd.notna(row.iloc[10]) else ''
            
            if pp_val and pp_val != 'nan' and pp_val != 'NAN' and partida_val and partida_val != '0000n':
                datos.append({'PP': pp_val, 'PARTIDA': partida_val})
    else:
        # Formato con columnas nombradas
        df_cols = pd.read_excel(archivo, dtype=str)
        col_pp = None
        col_partida = None
        
        for c in df_cols.columns:
            cu = str(c).upper()
            if 'PP' in cu or 'PROGRAMA' in cu:
                col_pp = c
            if 'PARTIDA' in cu or 'OBJETO' in cu:
                col_partida = c
        
        if col_pp and col_partida:
            for _, row in df_cols.iterrows():
                pp_val = str(row[col_pp]).strip().upper() if pd.notna(row[col_pp]) else ''
                partida_val = str(row[col_partida]).strip().zfill(5) if pd.notna(row[col_partida]) else ''
                if pp_val and pp_val != 'nan' and pp_val != 'NAN':
                    datos.append({'PP': pp_val, 'PARTIDA': partida_val})
    
    return datos


def validar_registros(datos, partidas_por_pp):
    """Valida los registros contra el catálogo."""
    resultados = []
    
    for d in datos:
        pp = d['PP']
        partida = d['PARTIDA']
        
        if pp not in partidas_por_pp:
            resultados.append({
                'PP': pp, 
                'PARTIDA': partida, 
                'VÁLIDO': 'NO', 
                'MOTIVO': f'Pp {pp} no existe en catálogo'
            })
        elif partida in partidas_por_pp[pp]:
            resultados.append({
                'PP': pp, 
                'PARTIDA': partida, 
                'VÁLIDO': 'SÍ', 
                'MOTIVO': ''
            })
        else:
            resultados.append({
                'PP': pp, 
                'PARTIDA': partida, 
                'VÁLIDO': 'NO', 
                'MOTIVO': 'Partida no autorizada para este Pp'
            })
    
    return resultados


def generar_excel_resultados(resultados):
    """Genera archivo Excel con resultados formateados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Validación"
    
    # Estilos
    header_fill = PatternFill(start_color='6B1D3D', end_color='6B1D3D', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    si_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    no_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    center = Alignment(horizontal='center', vertical='center')
    
    # Encabezados
    headers = ['PP', 'PARTIDA', 'VÁLIDO', 'MOTIVO']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    
    # Datos
    for i, r in enumerate(resultados, 2):
        ws.cell(row=i, column=1, value=r['PP']).border = border
        ws.cell(row=i, column=2, value=r['PARTIDA']).border = border
        
        cell_valido = ws.cell(row=i, column=3, value=r['VÁLIDO'])
        cell_valido.border = border
        cell_valido.alignment = center
        cell_valido.fill = si_fill if r['VÁLIDO'] == 'SÍ' else no_fill
        
        ws.cell(row=i, column=4, value=r['MOTIVO']).border = border
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    
    # Guardar a BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ══════════════════════════════════════════════════════════════════════════════
# INTERFAZ PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

# Header
st.markdown("""
<div class="main-header">
    <h1>✓ Validador Pp - Partida Específica</h1>
    <p>Verifica combinaciones de Programa Presupuestario y Partida según el catálogo oficial</p>
</div>
""", unsafe_allow_html=True)

# Sidebar - Cargar catálogo
with st.sidebar:
    st.markdown("### 📁 Catálogo Base")
    st.caption("Sube el archivo `Pp_-_Partida_Especifica_2026.xlsx`")
    
    catalogo_file = st.file_uploader(
        "Catálogo Pp-Partida",
        type=['xlsx', 'xls'],
        key="catalogo",
        label_visibility="collapsed"
    )
    
    if catalogo_file:
        partidas_por_pp = cargar_catalogo(catalogo_file)
        st.success(f"✓ Catálogo cargado")
        
        st.markdown("---")
        st.markdown("### 📊 Estadísticas")
        st.metric("Programas (Pp)", len(partidas_por_pp))
        st.metric("Total partidas", sum(len(v) for v in partidas_por_pp.values()))
        
        st.markdown("---")
        st.markdown("### 📋 Pps disponibles")
        pps_lista = sorted(partidas_por_pp.keys())
        st.text_area(
            "Lista de Pps",
            value=", ".join(pps_lista),
            height=150,
            label_visibility="collapsed"
        )

# Contenido principal - Tabs
if 'partidas_por_pp' not in dir() or not catalogo_file:
    st.info("👈 **Primero sube el catálogo** `Pp_-_Partida_Especifica_2026.xlsx` en la barra lateral")
else:
    tab1, tab2, tab3 = st.tabs([
        "🔍 Consulta Individual",
        "📋 Validación Masiva",
        "📖 Explorar Catálogo"
    ])
    
    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1: CONSULTA INDIVIDUAL
    # ══════════════════════════════════════════════════════════════════════════
    with tab1:
        st.markdown("#### Validar una combinación Pp + Partida")
        
        col1, col2, col3 = st.columns([2, 2, 1])
        
        with col1:
            pp_input = st.text_input(
                "Programa Presupuestario (Pp)",
                placeholder="Ej: S263, K017, E009",
                max_chars=10
            ).upper().strip()
        
        with col2:
            partida_input = st.text_input(
                "Partida específica",
                placeholder="Ej: 33104, 52301",
                max_chars=5
            ).strip().zfill(5) if st.session_state.get('partida_input_value') else ""
            partida_input = st.text_input(
                "Partida específica",
                placeholder="Ej: 33104, 52301",
                max_chars=5,
                key="partida_input_value"
            ).strip()
        
        with col3:
            st.markdown("<br>", unsafe_allow_html=True)
            validar_btn = st.button("Validar", type="primary", use_container_width=True)
        
        if validar_btn and pp_input:
            partida_check = partida_input.zfill(5) if partida_input else ""
            
            if pp_input not in partidas_por_pp:
                st.markdown(f"""
                <div class="result-invalid">
                    <strong>❌ Pp no encontrado</strong><br>
                    El programa <code>{pp_input}</code> no existe en el catálogo.
                </div>
                """, unsafe_allow_html=True)
                
                # Sugerencias
                similares = [p for p in partidas_por_pp.keys() if pp_input[0] in p][:5]
                if similares:
                    st.caption(f"¿Quisiste decir?: {', '.join(similares)}")
            
            elif not partida_check or partida_check == "00000":
                # Solo mostrar partidas del Pp
                partidas = sorted(partidas_por_pp[pp_input])
                st.success(f"✓ **Pp {pp_input}** tiene **{len(partidas)}** partidas válidas")
                
                # Agrupar por capítulo
                capitulos = {}
                for p in partidas:
                    cap = p[0]
                    if cap not in capitulos:
                        capitulos[cap] = []
                    capitulos[cap].append(p)
                
                for cap in sorted(capitulos.keys()):
                    with st.expander(f"Capítulo {cap}000 ({len(capitulos[cap])} partidas)"):
                        chips = " ".join([f"`{p}`" for p in capitulos[cap]])
                        st.markdown(chips)
            
            elif partida_check in partidas_por_pp[pp_input]:
                st.markdown(f"""
                <div class="result-valid">
                    <strong>✅ VÁLIDO</strong><br>
                    La partida <code>{partida_check}</code> está autorizada para el Pp <code>{pp_input}</code>
                </div>
                """, unsafe_allow_html=True)
            
            else:
                st.markdown(f"""
                <div class="result-invalid">
                    <strong>❌ NO VÁLIDO</strong><br>
                    La partida <code>{partida_check}</code> <strong>no</strong> está autorizada para el Pp <code>{pp_input}</code>
                </div>
                """, unsafe_allow_html=True)
                
                # Mostrar alternativas del mismo capítulo
                cap = partida_check[0]
                similares = sorted([p for p in partidas_por_pp[pp_input] if p[0] == cap])
                if similares:
                    st.caption(f"Partidas válidas del capítulo {cap}000:")
                    st.code(", ".join(similares[:20]))
    
    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2: VALIDACIÓN MASIVA
    # ══════════════════════════════════════════════════════════════════════════
    with tab2:
        st.markdown("#### Validar múltiples registros desde archivo")
        st.caption("Soporta formato PIPP o archivos con columnas Pp/Partida")
        
        archivo_validar = st.file_uploader(
            "Archivo con claves a validar",
            type=['xlsx', 'xls'],
            key="validar"
        )
        
        if archivo_validar:
            with st.spinner("Procesando archivo..."):
                datos = procesar_archivo_validacion(archivo_validar, partidas_por_pp)
            
            if not datos:
                st.error("No se encontraron registros válidos en el archivo")
                st.info("El archivo debe tener formato PIPP o columnas nombradas como Pp/Programa y Partida/Objeto")
            else:
                st.info(f"📋 **{len(datos)}** registros encontrados")
                
                if st.button("🔍 Validar registros", type="primary"):
                    resultados = validar_registros(datos, partidas_por_pp)
                    
                    # Estadísticas
                    validos = sum(1 for r in resultados if r['VÁLIDO'] == 'SÍ')
                    invalidos = len(resultados) - validos
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.markdown(f"""
                        <div class="stat-card">
                            <div class="stat-number">{len(resultados)}</div>
                            <div class="stat-label">Total registros</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col2:
                        st.markdown(f"""
                        <div class="stat-card success">
                            <div class="stat-number" style="color: #2E7D32">{validos}</div>
                            <div class="stat-label">Válidos ✓</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col3:
                        st.markdown(f"""
                        <div class="stat-card error">
                            <div class="stat-number" style="color: #C62828">{invalidos}</div>
                            <div class="stat-label">Con errores ✗</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # Tabla de resultados
                    st.markdown("---")
                    st.markdown("##### Detalle de validación")
                    
                    df_resultados = pd.DataFrame(resultados)
                    
                    # Colorear
                    def highlight_valid(row):
                        if row['VÁLIDO'] == 'SÍ':
                            return ['background-color: #E8F5E9'] * len(row)
                        return ['background-color: #FFEBEE'] * len(row)
                    
                    st.dataframe(
                        df_resultados.style.apply(highlight_valid, axis=1),
                        use_container_width=True,
                        height=400
                    )
                    
                    # Botón de descarga
                    excel_output = generar_excel_resultados(resultados)
                    st.download_button(
                        label="📥 Descargar resultados (.xlsx)",
                        data=excel_output,
                        file_name="Validacion_Pp_Partida.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    
    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3: EXPLORAR CATÁLOGO
    # ══════════════════════════════════════════════════════════════════════════
    with tab3:
        st.markdown("#### Explorar catálogo completo")
        
        # Selector de Pp
        pp_seleccionado = st.selectbox(
            "Selecciona un Programa Presupuestario",
            options=[""] + sorted(partidas_por_pp.keys()),
            format_func=lambda x: f"{x} ({len(partidas_por_pp.get(x, []))} partidas)" if x else "-- Seleccionar --"
        )
        
        if pp_seleccionado:
            partidas = sorted(partidas_por_pp[pp_seleccionado])
            
            st.success(f"**{pp_seleccionado}** tiene **{len(partidas)}** partidas autorizadas")
            
            # Agrupar por capítulo
            capitulos = {}
            for p in partidas:
                cap = p[0]
                if cap not in capitulos:
                    capitulos[cap] = []
                capitulos[cap].append(p)
            
            # Mostrar por capítulo
            for cap in sorted(capitulos.keys()):
                nombre_cap = {
                    '1': 'Servicios Personales',
                    '2': 'Materiales y Suministros',
                    '3': 'Servicios Generales',
                    '4': 'Transferencias',
                    '5': 'Bienes Muebles',
                    '6': 'Inversión Pública',
                    '7': 'Inversiones Financieras',
                    '8': 'Participaciones',
                    '9': 'Deuda Pública'
                }.get(cap, '')
                
                with st.expander(f"**Capítulo {cap}000** - {nombre_cap} ({len(capitulos[cap])} partidas)", expanded=True):
                    # Mostrar en grid
                    cols = st.columns(6)
                    for i, partida in enumerate(capitulos[cap]):
                        cols[i % 6].code(partida)


# ══════════════════════════════════════════════════════════════════════════════
# FOOTER
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("---")
st.caption("Validador Pp-Partida | SADER - Secretaría de Agricultura y Desarrollo Rural")
